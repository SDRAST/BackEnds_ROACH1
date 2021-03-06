# -*- coding: utf-8 -*-
import logging
import os
import Pyro5.api

from openpyxl import load_workbook
from openpyxl.reader.excel import InvalidFileException
import support.excel

module_logger = logging.getLogger(__name__)

modulepath = os.path.dirname(os.path.abspath(__file__))
module_logger.debug("path to this module: %s", modulepath)
paramfile = "model_params.xlsx"
module_logger.debug("model parameter file: %s", paramfile)

@Pyro5.api.expose
class FirmwareServer():
  """
  Serves information about firmware and their boffiles.

  Public attributes::
    firmware_wb - Excel spreadsheet with data on firmware
    logger      - class logger
    paramfile   - name of Excel spreadsheet
    parampath   - path to Excel spreadsheet
    param_ws    - 'Parameters' sheet
    register    - dict of dicts of register data
    sheetnames  - sheet names in Excel spreadsheet
  """
  def __init__(self, parampath=modulepath, paramfile=paramfile):
    """
    Create an instance of FirmwareServer()
    """
    self.parampath = parampath
    self.paramfile = paramfile
    self.logger = logging.getLogger(module_logger.name+".FirmwareServer")
    self.logger.debug("__init__: initialized")
    self._open_parameter_spreadsheet()

  def _open_parameter_spreadsheet(self):
    """
    Get the firmware summary worksheet
    """
    # self.logger.debug("_open_parameter_spreadsheet: for %s",
    #   self.parampath+self.paramfile)
    self.logger.debug("_open_parameter_spreadsheet: for {}".format(
      os.path.join(self.parampath,self.paramfile)
    ))
    try:
    #   self.firmware_wb = load_workbook(self.parampath+self.paramfile)
      self.firmware_wb = support.excel.load_workbook(os.path.join(self.parampath,self.paramfile))
    except IOError as details:
      self.logger.error(
      "_open_parameter_spreadsheet: loading spreadsheet failed with IO error.",
                        exc_info=True)
      raise IOError
    except InvalidFileException:
      self.logger.error(
      "_open_parameter_spreadsheet: .reader.excel doesn't like this file.",
                        exc_info=True)
      raise InvalidFileException
    except AttributeError:
      self.logger.error(
                        "_open_parameter_spreadsheet: attribute error.",
                        exc_info=True)
      raise AttributeError
    self.sheet_names = self.firmware_wb.get_sheet_names()
    self.logger.debug("_open_parameter_spreadsheet: sheet names: %s",
                      str(self.sheet_names))
    self.param_ws = self.firmware_wb.get_sheet_by_name('Parameters')
    #column_names = get_column_names(self.param_ws)
    #self.logger.debug("_open_parameter_spreadsheet: columns found:")
    #for name in column_names.keys():
    #  if column_names[name]:
    #    self.logger.debug("_open_parameter_spreadsheet: %s: %s",
    #                      name, column_names[name])

  def get_keys(self,sheet=''):
    """
    Get the keys for the rows of the designated sheet

    @param sheet : 'Parameters' sheet if blank
    @type  sheet : str

    @return: list of entries in the 'key' or 'Register' column
    """
    if sheet == '':
      sheet = self.param_ws
      self.logger.debug('get_keys: checking sheet %s', sheet)
      keys = support.excel.get_column(sheet,'key')
    else:
      sheet = self.firmware_wb.get_sheet_by_name(sheet)
      keys = support.excel.get_column(sheet,'Register')
    return keys

  def firmware_summary(self,key):
    """
    Get the summary data for the designated firmware.

    @param key : item from first column of 'Parameters' sheet

    @return: dict with data from "Parameters" sheet row
    """
    summary = {}
    self.logger.debug("firmware_summary: for %s", self.param_ws)
    # Get the column names
    # Create a dictionary keyed with column names
    col_numbers = {}
    number = 1
    for col in self.param_ws.iter_cols(1, self.param_ws.max_column):
      col_numbers[col[0].value] = number
      number += 1
    self.logger.debug("firmware summary: column name dict: %s", col_numbers)
    # get the firmware key column
    col_num = col_numbers['key']
    self.logger.debug("firmware_summary: column for 'key' is %s", col_num)
    # create a dictionary keyed with row names
    row_numbers = {}
    number = 1
    for row in self.param_ws.iter_rows(1, self.param_ws.max_row):
      row_numbers[row[0].value] = number
      number += 1
    self.logger.debug("firmware_summary: row name dict: %s", row_numbers)
    # get selected firmware row
    row_number = row_numbers[key]
    # now generate the summary
    summary['row'] = row_number
    bitstream = self.param_ws.cell(row=row_number, 
                                   column=col_numbers['bitstream']).value
    summary['bitstream'] = bitstream
    n_chans = int(self.param_ws.cell(row=row_number,
                                 column=col_numbers['nchans']).value)
    summary['nchans'] = n_chans
    n_par_streams = int(self.param_ws.cell(row=row_number,
                                     column=col_numbers['n_par_streams']).value)
    summary['n_par_streams'] = n_par_streams
    fft_shift = int(self.param_ws.cell(row=row_number,
                                       column=col_numbers['fft_shift']).value)
    if type(fft_shift) == str or type(fft_shift) == str:
      fft_shift = int(fft_shift[2:],2)
    summary["fft_shift"] = fft_shift
    summary["desired_rf_level"] =self.param_ws.cell(row=row_number,
                                   column=col_numbers['desired_rf_level']).value
    summary["spectrum_bits"] = self.param_ws.cell(row=row_number,
                                      column=col_numbers['spectrum_bits']).value
    bandwidth = self.param_ws.cell(row=row_number,
                                   column=col_numbers['design bandwidth']).value
    summary['bandwidth'] = bandwidth
    interleaved = self.param_ws.cell(row=row_number,
                                        column=col_numbers['interleaved']).value
    if interleaved:
      clock = bandwidth
    else:
      clock = 2*bandwidth
    summary['clock'] = clock
    sys_clock = (1+interleaved)*clock/n_par_streams
    summary['sys_clock'] = sys_clock
    ADC_input_str = self.param_ws.cell(row=row_number,
                                       column=col_numbers['ADC inputs']).value
    [ADC0str,ADC1str] = ADC_input_str.split(';')
    # There must always be an ADC0
    [ADC0in0str,ADC0in1str] = ADC0str.split(',')
    ADC_inputs = {0: [int(ADC0in0str)]}
    if ADC0in1str:
      ADC_inputs[0].append(int(ADC0in1str))
    [ADC1in0str,ADC1in1str] = ADC1str.split(',')
    if ADC1in0str:
      ADC_inputs[1] = [int(ADC1in0str)]
      if ADC1in1str:
        ADC_inputs[1].append(int(ADC1in1str))
    summary['ADC inputs'] = ADC_inputs
    ADC_type_str = self.param_ws.cell(row=row_number,
                                      column=col_numbers['adc_type']).value
    ADC_type_list = ADC_type_str.split(',')
    ADC_type = {0: ADC_type_list[0]}
    if len(ADC_type_list) == 2:
      ADC_type[1] = ADC_type_list[1]
    summary['ADC types'] = ADC_type
    for index in range(4):
      gbe = 'gbe'+str(index)
      self.logger.debug("firmware_summary: processing %s", gbe)
      column_name = gbe +' MAC'
      self.logger.debug("firmware_summary: checking %s", column_name)
      gbe_MAC = self.param_ws.cell(row=row_number,
                                   column=col_numbers[column_name]).value
      self.logger.debug("firmware_summary: MAC is %s", gbe_MAC)
      if gbe_MAC:
        summary[gbe+' MAC'] = gbe_MAC
        summary[gbe+' IP'] = self.param_ws.cell(row=row_number,
                                           column=col_numbers[gbe +' IP']).value
    self.logger.debug("firmware_summary so far: %s", summary)
    return summary

  def parse_registers(self,sheetname):
    """
    Get the register functions

    @param sheetname : name of firmware data sheet
    @type  sheetname : str

    @return: dict of dicts with register data
    """
    sheet = self.firmware_wb.get_sheet_by_name(sheetname)
    self.logger.debug("parse_registers: parsing sheet %s", sheet)
    self.register = {}
    for row in sheet.rows[1:]:
      # this means that index 0 is row 2, because we start with the second row
      rowindex = sheet.rows[1:].index(row)
      bits = str(row[2].value)
      if row[0].value:
        reg_ID = str(row[0].value)
        #self.logger.debug("parse_registers: processing '%s'",reg_ID)
        self.register[reg_ID] = {'direction': str(row[1].value)}
        #self.logger.debug("parse_registers: bits = '%s'",bits)
        if bits:
          if bits != "31:0":
            self.register[reg_ID]['bits'] = {bits:{'name':str(row[3].value)}}
            self.register[reg_ID]['bits'][bits]['purpose'] = str(row[4].value)
            if row[5].value:
              self.register[reg_ID]['bits'][bits]['values'] = str(row[5].value)
            prev_ID = reg_ID
          else:
            self.register[reg_ID]['bits'] = {bits: None}
      else:
        #self.logger.debug("parse_registers: row %s column 0 has no value", rowindex+2)
        if bits != 'None':
          self.register[prev_ID]['bits'][bits] = {'name':str(row[3].value)}
          self.register[prev_ID]['bits'][bits]['purpose'] = str(row[4].value)
          if row[5].value:
            self.register[prev_ID]['bits'][bits]['values'] = \
              str(row[5].value)
    return self.register
